import json
import os
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from zoneinfo import ZoneInfo

PARAGUAY_TZ = ZoneInfo("America/Asuncion")
STORAGE_PATH = Path(__file__).parent
HISTORY_FILE = str(STORAGE_PATH / "fbox_history.json")
ALERTS_HISTORY_FILE = str(STORAGE_PATH / "fbox_alerts_history.json")

def generate_excel():
    """Genera reporte Excel con historial de estados y alertas"""
    
    # Cargar datos
    try:
        with open(HISTORY_FILE, 'r') as f:
            history = json.load(f)
    except:
        print("❌ No se encontró historial de estados")
        return
    
    try:
        with open(ALERTS_HISTORY_FILE, 'r', encoding='utf-8') as f:
            alerts_history = json.load(f)
    except:
        alerts_history = []
    
    if not history:
        print("❌ No hay datos en el historial")
        return
    
    # Crear workbook
    wb = Workbook()
    wb.remove(wb.active)
    
    # ============ HOJA 1: RESUMEN ============
    ws_summary = wb.create_sheet("Resumen", 0)
    ws_summary['A1'] = "📊 REPORTE SEMANAL FBOX"
    ws_summary['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws_summary['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws_summary.merge_cells('A1:B1')
    
    ws_summary['A3'] = f"Período: {history[0].get('timestamp', '').split('T')[0]} a {history[-1].get('timestamp', '').split('T')[0]}"
    ws_summary['A4'] = f"Generado: {datetime.now(PARAGUAY_TZ).strftime('%Y-%m-%d %H:%M:%S')}"
    ws_summary['A5'] = f"Total de registros: {len(history)}"
    ws_summary['A6'] = f"Total de alertas: {len(alerts_history)}"
    
    ws_summary.column_dimensions['A'].width = 50
    
    # ============ HOJA 2: HISTORIAL DE ESTADOS ============
    ws_history = wb.create_sheet("Historial")
    
    # Headers
    headers = ["Timestamp", "Contenedor", "Status", "Temp Aceite (°C)", "Temp Contenedor (°C)", 
               "Mineros Online", "Mineros Offline", "Hashrate (PH/s)", "Potencia (kW)"]
    ws_history.append(headers)
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for cell in ws_history[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
    
    # Datos
    for record in history:
        timestamp = record.get("timestamp", "")
        data = record.get("data", {})
        
        for container_name, container_data in data.items():
            status = "🟢 ONLINE" if container_data.get("code") == 1 else "🔴 OFFLINE"
            row = [
                timestamp,
                container_name,
                status,
                container_data.get("oil_temp"),
                container_data.get("container_temp"),
                container_data.get("miner_online"),
                container_data.get("miner_offline"),
                container_data.get("hashrate_ph"),
                container_data.get("power_kw")
            ]
            ws_history.append(row)
            
            # Aplicar estilos a la fila
            for cell in ws_history[ws_history.max_row]:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")
    
    # Ajustar ancho de columnas
    ws_history.column_dimensions['A'].width = 20
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws_history.column_dimensions[col].width = 15
    
    # ============ HOJA 3: ALERTAS ============
    if alerts_history:
        ws_alerts = wb.create_sheet("Alertas")
        
        # Headers
        alert_headers = ["Timestamp", "Alerta"]
        ws_alerts.append(alert_headers)
        
        for cell in ws_alerts[1]:
            cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.border = thin_border
        
        # Datos de alertas
        for record in alerts_history:
            timestamp = record.get("timestamp", "")
            alerts = record.get("alerts", [])
            
            for alert in alerts:
                row = [timestamp, alert]
                ws_alerts.append(row)
                
                for cell in ws_alerts[ws_alerts.max_row]:
                    cell.border = thin_border
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
        
        ws_alerts.column_dimensions['A'].width = 20
        ws_alerts.column_dimensions['B'].width = 80
    
    # Guardar archivo
    timestamp = datetime.now(PARAGUAY_TZ).strftime("%Y%m%d_%H%M%S")
    filename = f"FBOX_Report_{timestamp}.xlsx"
    filepath = STORAGE_PATH / filename
    
    wb.save(filepath)
    print(f"✅ Reporte Excel generado: {filepath}")
    return filepath

if __name__ == "__main__":
    generate_excel()
